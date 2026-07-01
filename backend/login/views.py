import json
import os
import uuid
import openpyxl
import logging
from datetime import date, datetime
from pathlib import Path

from django.conf import settings
from django.db import DatabaseError
from django.db import connection
from django.http import JsonResponse
from django.utils.text import get_valid_filename
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.core.exceptions import ValidationError
from .validators import validate_uploaded_excel, run_validation_checklist, analyze_and_generate_flags
from rest_framework import status as drf_status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from .auth_service import auth_service_method, is_auth_bypass_user
from .db import check_database_connection, get_payroll_connection
from .models import EmployeeMaster, RoleMaster
from .serializers import EmployeeMasterSerializer, FreezeIntervalSerializer, RoleMasterSerializer

VALID_DOMAINS = {
    'ASG',
    'Cadomain',
    'EVEREST',
    'Medassistgroup',
    'Medplans',
    'MLD.Cadomain.Local',
    'TKGCorp',
    'to.ccrs.com',
}

LOGIN_FAILURE_MESSAGE = 'Login Failure, Did you forget or misspell your ID or password?'
GEO_OPTIONS = ('India', 'Philippines')
DOCUMENT_TYPE_MAP = {
    'payroll': 'PRL',
    'prl': 'PRL',
    'irefer': 'IRF',
    'i refer': 'IRF',
    'irf': 'IRF',
    'transport': 'TDCT',
    'transportdeduction': 'TDCT',
    'transport deduction': 'TDCT',
    'tdct': 'TDCT',
}
DOCUMENT_LABELS = {
    'PRL': 'Payroll',
    'IRF': 'IRefer',
    'TDCT': 'TransportDeduction',
}
DOCUMENT_FOLDERS = {
    'PRL': {'India': 'PayrollPANIndia', 'Philippines': 'PayrollPHP'},
    'IRF': {'India': 'IRefer', 'Philippines': 'IReferPHP'},
    'TDCT': {'India': 'TransportDeduction', 'Philippines': 'TransportDeduction'},
}


def employee_to_dict(employee):
    return {
        'EmployeeCode': employee.employee_code,
        'DOMAIN_ID': employee.domain_id,
        'EmployeeName': employee.employee_name,
        'Role': employee.role,
        'Grade': employee.grade,
        'SupervisorCode': employee.supervisor_code,
        'EmployeeEmailId': employee.employee_email_id,
        'EmpGeo': employee.emp_geo,
    }


def get_employee_by_domain_id(domain_id):
    employee = (
        EmployeeMaster.objects.filter(domain_id=domain_id, deleted_status=False)
        .only(
            'employee_code',
            'domain_id',
            'employee_name',
            'role',
            'grade',
            'supervisor_code',
            'employee_email_id',
            'emp_geo',
        )
        .first()
    )
    if employee is None:
        return None
    return employee_to_dict(employee)


def build_user_payload(domain_name, employee):
    role_id = employee.get('Role')
    
    if str(role_id) == '3':
        redirect_url = '/approver'
    elif str(role_id) == '2':
        redirect_url = '/admin-report'
    else:
        redirect_url = '/home'

    return {
        'domainName': domain_name,
        'domainId': employee.get('DOMAIN_ID'),
        'employeeCode': employee.get('EmployeeCode'),
        'employeeName': employee.get('EmployeeName'),
        'role': role_id,
        'grade': employee.get('Grade'),
        'supervisor': employee.get('SupervisorCode'),
        'emailId': employee.get('EmployeeEmailId'),
        'empGeo': employee.get('EmpGeo'),
        'loginType': 'LDAP',
        'redirectUrl': redirect_url,
    }


def store_login_session(request, employee):
    request.session['LoginID'] = employee.get('EmployeeCode')
    request.session['EmployeeName'] = employee.get('EmployeeName')
    request.session['Supervisor'] = employee.get('SupervisorCode')
    request.session['RoleID'] = employee.get('Role')
    request.session['Grade'] = employee.get('Grade')
    request.session['DomainId'] = employee.get('DOMAIN_ID')
    request.session['EmailId'] = employee.get('EmployeeEmailId')
    request.session['EmpGeo'] = employee.get('EmpGeo')
    request.session['LoginType'] = 'LDAP'
    request.session['RootDir'] = settings.ROOT_DIR


def login_failure(message=LOGIN_FAILURE_MESSAGE, status=401):
    return JsonResponse({'success': False, 'message': message}, status=status)


def api_index(request):
    return JsonResponse(
        {
            'service': 'FirstPay Phase 2 login API',
            'loginEndpoint': '/api/login/',
            'logoutEndpoint': '/api/logout/',
            'dbStatusEndpoint': '/api/db-status/',
            'dashboardEndpoint': '/api/dashboard/',
            'adminReportEndpoint': '/api/admin-report/',
            'method': 'POST',
            'database': settings.PAYROLL_SQL_DATABASE,
            'server': settings.PAYROLL_SQL_SERVER,
            'authMode': (
                'Integrated Security'
                if settings.PAYROLL_SQL_TRUSTED_CONNECTION.lower() == 'yes'
                else 'SQL Login'
            ),
        }
    )


def sample_upload_row(row_id, user_id, document_type, uploaded_date, status, remarks):
    return {
        'id': row_id,
        'userId': user_id,
        'documentType': document_type,
        'uploadedDate': uploaded_date,
        'status': status,
        'remarks': remarks,
        'templateUrl': '#',
        'emailUrl': '#',
    }


def normalize_document_type(value):
    text = str(value or '').strip().lower()
    return DOCUMENT_TYPE_MAP.get(text, DOCUMENT_TYPE_MAP.get(text.replace(' ', ''), str(value or '').strip().upper()))


def safe_text(value, fallback=''):
    if value is None:
        return fallback
    return str(value)


def format_uploaded_date(value):
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y - %H:%M')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    return safe_text(value)


def row_value(row, *names, fallback=''):
    lowered = {str(key).lower(): value for key, value in row.items()}
    for name in names:
        val = None
        if name in row:
            val = row[name]
        else:
            val = lowered.get(str(name).lower())
        if val is not None and str(val).strip() != '' and str(val).strip().lower() != 'none':
            return val
    return fallback



def public_file_url(path):
    text = safe_text(path)
    if not text:
        return '#'
    if text.startswith(('http://', 'https://', '/')):
        return text
    return f'/{text.lstrip("./")}'


def upload_row_from_db(row, fallback_id, log_details=None):
    doc_code = normalize_document_type(row_value(row, 'DocumentType', 'TypeOfDoc', 'DocType'))
    log_id = int(row_value(row, 'LogId', 'MMLogId', 'Id', 'ID', fallback=fallback_id) or fallback_id)
    
    details = log_details.get(log_id, {}) if log_details else {}
    
    return {
        'id': log_id,
        'mergeId': row_value(row, 'MMLogId', 'MergeId', 'MMId', fallback=''),
        'userId': safe_text(row_value(row, 'UploadedByEmpNo', 'UploadedBy', 'Uploadedby', 'UserID', 'EmpNo')),
        'documentType': safe_text(row_value(row, 'DocumentType', 'TypeOfDoc', fallback=DOCUMENT_LABELS.get(doc_code, doc_code))),
        'documentCode': doc_code,
        'uploadedDate': format_uploaded_date(row_value(row, 'CreatedDate', 'UploadedDate', 'TimeStamp')),
        'status': safe_text(row_value(row, 'State', 'Status', fallback='Pending')),
        'stateId': int(row_value(row, 'StateId', 'MM_StateId', fallback=3) or 3),
        'remarks': safe_text(row_value(row, 'Remarks', 'Comments')),
        'templateUrl': public_file_url(row_value(row, 'S3Link', 'FilePath', 'TemplatePath')),
        'emailUrl': public_file_url(row_value(row, 'FilePathEmail', 'EmailPath')),
        'filePath': public_file_url(row_value(row, 'FilePath', 'S3Link', 'TemplatePath')),
        'approverName': details.get('approver', ''),
        'approvalDate': details.get('approvalDate', ''),
    }


def empty_upload_groups(include_consolidated=False):
    groups = {'payroll': [], 'irefer': [], 'transport': []}
    if include_consolidated:
        groups['consolidated'] = []
    return groups


def append_upload_rows(groups, rows, fallback_offset=0, log_details=None):
    for index, row in enumerate(rows, start=1):
        item = upload_row_from_db(row, fallback_offset + index, log_details)
        doc_code = item.get('documentCode')
        key = {'PRL': 'payroll', 'IRF': 'irefer', 'TDCT': 'transport'}.get(doc_code, 'payroll')
        groups.setdefault(key, []).append(item)


def execute_dataset_proc(proc_name, params):
    result_sets = []
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        assignments = ', '.join(f'{name}=?' for name, _ in params)
        cursor.execute(f'EXEC {proc_name} {assignments}', [value for _, value in params])
        while True:
            if cursor.description:
                columns = [column[0] for column in cursor.description]
                result_sets.append([dict(zip(columns, row)) for row in cursor.fetchall()])
            if not cursor.nextset():
                break
        return result_sets
    finally:
        conn.close()


def get_uploaded_files_payload(action, request, include_consolidated=False):
    login_id = get_session_login_id(request)
    domain_id = str(request.session.get('DomainId') or login_id)
    user_geo = get_session_geo(request)
    result_sets = execute_dataset_proc(
        'SP_GetAllUploadedFiles',
        [('@LoginId', domain_id), ('@Action', action), ('@UserGeo', user_geo)],
    )
    
    # Extract log IDs to fetch Approver and Approval/Rejection Date
    log_ids = []
    for r_set in result_sets:
        for r in r_set:
            lid = r.get('LogId') or r.get('MMLogId') or r.get('Id') or r.get('ID')
            if lid:
                log_ids.append(lid)
                
    log_details = {}
    if log_ids:
        log_ids_str = ', '.join(str(int(lid)) for lid in log_ids)
        conn = get_payroll_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(f"SELECT LogId, Approver, AdminApproverejectDate FROM tbl_UploadSuccessLog WHERE LogId IN ({log_ids_str})")
            for log_id, approver, app_date in cursor.fetchall():
                log_details[log_id] = {
                    'approver': approver or '',
                    'approvalDate': app_date.strftime("%d/%b/%Y") if app_date else ''
                }
        except Exception as e:
            logging.error(f"Error loading extra log columns: {e}")
        finally:
            conn.close()

    groups = empty_upload_groups(include_consolidated)
    if action == 'AdminSelect':
        names = ['payroll', 'irefer', 'transport', 'email', 'payroll', 'irefer', 'transport']
        for set_index, rows in enumerate(result_sets):
            if set_index < 3:
                append_upload_rows({names[set_index]: groups[names[set_index]]}, rows, set_index * 1000, log_details)
            elif include_consolidated and set_index >= 4:
                for row_index, row in enumerate(rows, start=1):
                    groups['consolidated'].append(upload_row_from_db(row, 4000 + set_index * 100 + row_index, log_details))
    else:
        for set_index, rows in enumerate(result_sets[:3]):
            append_upload_rows(groups, rows, set_index * 1000, log_details)
    return groups


def build_dashboard_fixture():
    return {
        'payroll': [
            sample_upload_row(1, '1038798', 'Payroll', '29/06/2025 - 05:08', 'Uploaded', 'Ready for validation'),
            sample_upload_row(2, '1038812', 'Payroll', '30/06/2025 - 10:30', 'Under Review', 'Pending manager review'),
        ],
        'irefer': [
            sample_upload_row(3, '1038877', 'IRefer', '30/06/2025 - 11:20', 'Under Review', 'Pending validation'),
        ],
        'transport': [
            sample_upload_row(4, '1038890', 'TransportDeduction', '01/07/2025 - 09:45', 'Uploaded', 'Processed'),
        ],
    }


def get_session_geo(request):
    return str(request.session.get('EmpGeo') or 'India').strip() or 'India'


def build_templates_fixture(emp_geo):
    templates = [
        {
            'id': 1,
            'code': 'PRL',
            'imgSrc': '/assets/images/templates/payroll.png',
            'title': 'Payroll',
            'desc': (
                'Use this official Payroll Automation Template to fill in all required employee payroll data. '
                'The downloadable format (.xls/.xlsx) includes predefined columns such as Employee ID, Name, '
                'Basic Pay, Allowances, Deductions, and Net Pay. '
            ),
            'fileLink': {
                'India': '/DownloadTemplates/PayrollPANIndia/PayrollFile.xlsx',
                'Philippines': '/DownloadTemplates/PayrollPhp/PayrollPhilippines.xlsx',
            },
        },
        {
            'id': 2,
            'code': 'TDCT',
            'imgSrc': '/assets/images/templates/transport.png',
            'title': 'Transport Deduction',
            'desc': (
                'It allows you to record and submit employee transport-related deductions for the current payroll '
                'cycle. This ensures that any commute or company transport cost recovery is accurately reflected '
                'in the final salary disbursement. '
            ),
            'fileLink': {
                'India': '/DownloadTemplates/TransportDeduction/TransportDeductionFile.xlsx',
            },
        },
        {
            'id': 3,
            'code': 'IRF',
            'imgSrc': '/assets/images/templates/irefer.png',
            'title': 'IRefer',
            'desc': (
                'Use the standardized IRefer Template to upload referral-related information for payroll processing. '
                'This format ensures accurate capture of associate and referral details, along with the eligible '
                'payout amount. Only .xls or .xlsx formats are accepted. ocess.'
            ),
            'fileLink': {
                'India': '/DownloadTemplates/IRefer/IreferFile.xlsx',
                'Philippines': '/DownloadTemplates/IRefer/IreferFile.xlsx',
            },
        },
    ]
    if emp_geo == 'Philippines':
        templates = [template for template in templates if template['id'] != 2]
    return templates


def get_session_login_id(request):
    return str(request.session.get('LoginID') or request.session.get('DomainId') or 'system')


def coerce_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def freeze_interval_status(start_date, end_date, role_id=None):
    today = timezone.localdate()
    is_between = bool(start_date and end_date and start_date <= today <= end_date)
    return {
        'today': today.isoformat(),
        'isCurrentDateInRange': is_between,
        'isUploadScreenFreezed': is_between and str(role_id or '') == '1',
        'summaryClassName': 'divSummaryFreezeDate bg-success text-white' if is_between else 'divSummaryFreezeDate bg-danger text-white',
        'statusText': '(Upload Screen Freezed !)' if is_between else '(Upload-Screen Not Freeze.)',
    }


def execute_freeze_interval(start_date='', end_date='', login_id='system', action='GET-INTERVAL'):
    with connection.cursor() as cursor:
        cursor.execute(
            'EXEC SP_SetFreezeInterval @StartDt=%s, @EndDt=%s, @LoginId=%s, @Action=%s',
            [start_date, end_date, login_id, action],
        )
        description = cursor.description or []
        columns = [column[0] for column in description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()] if columns else []
    return rows


def build_freeze_interval_payload(row, request):
    start_date = coerce_date(row.get('StartDate') if row else None)
    end_date = coerce_date(row.get('EndDate') if row else None)
    status_payload = freeze_interval_status(start_date, end_date, request.session.get('RoleID'))
    return {
        'success': True,
        'data': {
            'startDate': start_date.isoformat() if start_date else '',
            'endDate': end_date.isoformat() if end_date else '',
            **status_payload,
        },
    }


@api_view(['GET', 'POST'])
def freeze_interval(request):
    login_id = get_session_login_id(request)

    if request.method == 'GET':
        rows = execute_freeze_interval('', '', login_id, 'GET-INTERVAL')
        return Response(build_freeze_interval_payload(rows[0] if rows else None, request))

    serializer = FreezeIntervalSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    start_date = serializer.validated_data['startDate']
    end_date = serializer.validated_data['endDate']
    rows = execute_freeze_interval(start_date.isoformat(), end_date.isoformat(), login_id, 'SET-INTERVAL')
    row = rows[0] if rows else {'StartDate': start_date, 'EndDate': end_date}
    return Response(build_freeze_interval_payload(row, request))


@api_view(['GET', 'POST'])
def users_collection(request):
    if request.method == 'GET':
        users = EmployeeMaster.objects.filter(deleted_status=False).order_by('employee_code', 'employee_name')
        serializer = EmployeeMasterSerializer(users, many=True)
        return Response({'success': True, 'totalEmployees': users.count(), 'data': serializer.data})

    serializer = EmployeeMasterSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    employee = serializer.save(
        deleted_status=False,
        created_by=get_session_login_id(request),
        created_date=timezone.now(),
    )
    return Response(
        {
            'success': True,
            'message': 'New employee created successfully!',
            'data': EmployeeMasterSerializer(employee).data,
        },
        status=drf_status.HTTP_201_CREATED,
    )


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
def users_detail(request, pk):
    try:
        employee = EmployeeMaster.objects.get(pk=pk, deleted_status=False)
    except EmployeeMaster.DoesNotExist:
        return Response({'success': False, 'message': 'Employee not found.'}, status=drf_status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response({'success': True, 'data': EmployeeMasterSerializer(employee).data})

    if request.method in ('PUT', 'PATCH'):
        serializer = EmployeeMasterSerializer(
            employee,
            data=request.data,
            partial=request.method == 'PATCH',
        )
        serializer.is_valid(raise_exception=True)
        employee = serializer.save(
            last_edited_by=get_session_login_id(request),
            last_edited_date=timezone.now(),
        )
        return Response(
            {
                'success': True,
                'message': 'Details Saved Successfully !',
                'data': EmployeeMasterSerializer(employee).data,
            }
        )

    employee.delete()
    return Response({'success': True, 'message': 'Record deleted successfully !'})


@api_view(['GET'])
def users_options(request):
    roles = list(RoleMaster.objects.order_by('role_id'))
    if not roles:
        roles = [
            RoleMaster(role_id=2, role_name='Admin'),
            RoleMaster(role_id=1, role_name='User'),
        ]

    return Response(
        {
            'success': True,
            'roles': RoleMasterSerializer(roles, many=True).data,
            'geos': [{'Value': geo, 'Text': geo} for geo in GEO_OPTIONS],
        }
    )


@require_GET
def templates_report(request):
    emp_geo = get_session_geo(request)
    return JsonResponse({'success': True, 'empGeo': emp_geo, 'data': build_templates_fixture(emp_geo)})


@require_GET
def dashboard_report(request):
    try:
        data = get_uploaded_files_payload('Select', request)
        return JsonResponse({'success': True, 'source': 'database', 'data': data})
    except Exception as exc:
        return JsonResponse({'success': True, 'source': 'fixture', 'warning': str(exc), 'data': build_dashboard_fixture()})


@require_GET
def admin_report(request):
    try:
        data = get_uploaded_files_payload('AdminSelect', request, include_consolidated=True)
        return JsonResponse({'success': True, 'source': 'database', 'data': data})
    except Exception as exc:
        data = build_dashboard_fixture()
        data['payroll'] = [
            sample_upload_row(101, '1038798', 'Payroll', '29/06/2025 - 05:08', 'Pending', 'Waiting for SPOC approval'),
            sample_upload_row(102, '1038812', 'Payroll', '30/06/2025 - 08:18', 'Approved', 'Approved by manager'),
        ]
        data['irefer'] = [
            sample_upload_row(201, '1038877', 'IRefer', '30/06/2025 - 11:20', 'Rejected', 'Incorrect amount'),
        ]
        data['transport'] = [
            sample_upload_row(301, '1038890', 'TransportDeduction', '01/07/2025 - 09:45', 'Pending', 'New request'),
        ]
        data['consolidated'] = [
            sample_upload_row(401, 'Admin', 'Payroll', '02/07/2025 - 16:10', 'Consolidated', 'Monthly payroll consolidated'),
        ]
        return JsonResponse({'success': True, 'source': 'fixture', 'warning': str(exc), 'data': data})


def save_uploaded_file(uploaded_file, folder_name, prefix):
    upload_root = Path(settings.BASE_DIR) / 'UploadFiles'
    target_dir = upload_root / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)
    original_name = get_valid_filename(uploaded_file.name)
    stored_name = f'{prefix}_{uuid.uuid4().hex[:10]}_{original_name}'
    target_path = target_dir / stored_name
    with target_path.open('wb+') as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
    relative_path = f'/UploadFiles/{folder_name}/{stored_name}'
    return relative_path, original_name


def inject_uploaded_by_column(file_path, domain_id):
    try:
        wb = openpyxl.load_workbook(file_path)
        modified = False
        for sheet in wb.worksheets:
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            last_col_idx = 0
            for r in range(1, max_row + 1):
                # find the last cell with value in this row
                for c_idx in range(max_col, 0, -1):
                    if sheet.cell(row=r, column=c_idx).value is not None:
                        if c_idx > last_col_idx:
                            last_col_idx = c_idx
                        break
            
            new_col = last_col_idx + 1
            sheet.cell(row=1, column=new_col, value="uploaded by")
            
            for r in range(2, max_row + 1):
                sheet.cell(row=r, column=new_col, value=domain_id)
                
            modified = True
            
        if modified:
            wb.save(file_path)
    except Exception as e:
        logging.error(f"Failed to inject uploaded by column into {file_path}: {e}")


@csrf_exempt
@require_POST
def submit_upload(request):
    template_file = request.FILES.get('templateFile')
    email_file = request.FILES.get('emailFile')
    doc_code = normalize_document_type(request.POST.get('documentType'))
    user_geo = request.POST.get('usergeo') or get_session_geo(request)

    if not doc_code or doc_code not in DOCUMENT_LABELS:
        return JsonResponse({'success': False, 'message': 'select your option !'}, status=400)
    if template_file is None:
        return JsonResponse({'success': False, 'message': 'Required Fields Check: Verifies that a document type is selected and template file is attached.'}, status=400)

    extension = os.path.splitext(template_file.name)[1].lower()
    if extension not in ('.xls', '.xlsx'):
        return JsonResponse({'success': False, 'message': 'Upload failure ! only .xls and .xlsx formats allowed !'}, status=400)

    if email_file is not None:
        email_extension = os.path.splitext(email_file.name)[1].lower()
        if email_extension not in ('.msg', '.eml'):
            return JsonResponse({'success': False, 'message': 'Only .msg / .eml message document for email approval are allowed.'}, status=400)

    # Perform structural and content validations on the Excel template
    try:
        validate_uploaded_excel(template_file, doc_code, user_geo)
    except ValidationError as e:
        return JsonResponse({'success': False, 'message': e.message}, status=400)

    folder_name = DOCUMENT_FOLDERS[doc_code].get(user_geo, DOCUMENT_FOLDERS[doc_code]['India'])
    login_id = get_session_login_id(request)
    domain_id = str(request.session.get('DomainId') or login_id)
    prefix = f'{domain_id}_{timezone.now().strftime("%d%m%Y%H%M%S")}_{doc_code}'
    file_path, original_name = save_uploaded_file(template_file, folder_name, prefix)
    
    # Inject "uploaded by" column
    abs_file_path = Path(settings.BASE_DIR) / file_path.lstrip('/')
    inject_uploaded_by_column(str(abs_file_path), domain_id)
    
    email_path = ''
    if email_file is not None:
        email_path, _ = save_uploaded_file(email_file, 'Emails', f'{domain_id}_{timezone.now().strftime("%d%m%Y%H%M%S")}_EMAIL')

    # For User (Role 1), only 1 request is allowed per month per geography.
    # If a previous active request exists, soft delete it (IsDeleted = 1).
    role_id = str(request.session.get('RoleID') or '1')
    if role_id == '1':
        now = timezone.now()
        current_year = now.year
        current_month = now.month
        try:
            conn = get_payroll_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE tbl_UploadSuccessLog 
                SET IsDeleted = 1, ModifiedDate = GETDATE(), ModifiedBy = ?
                WHERE Uploadedby = ? AND DocumentType = ? AND UserGeo = ? AND (IsDeleted = 0 OR IsDeleted IS NULL)
                AND YEAR(CreatedDate) = ? AND MONTH(CreatedDate) = ?
            """, [domain_id, domain_id, doc_code, user_geo, current_year, current_month])
            conn.commit()
            conn.close()
        except Exception as e:
            logging.error(f"Error soft deleting previous request: {e}")

    try:
        rows = execute_dataset_proc(
            'SP_InsertSuccessLog',
            [
                ('@Type', doc_code),
                ('@StrFilePath', file_path),
                ('@StrFileName', original_name),
                ('@UploadBy', domain_id),
                ('@Action', 'Insert'),
                ('@EmpNo', login_id),
                ('@EmailStrFilePath', email_path),
                ('@UserGeo', user_geo),
            ],
        )
        request_id = ''
        if rows and rows[0]:
            request_id = safe_text(row_value(rows[0][0], 'New_ID', 'RequestId', 'Id'))
            if request_id:
                try:
                    import re
                    approver_label = request.POST.get('approver')
                    approver_email = ''
                    if approver_label:
                        match = re.search(r'[\w\.-]+@[\w\.-]+', approver_label)
                        if match:
                            approver_email = match.group(0).strip().lower()
                    if approver_email:
                        approver_login = ''
                        emp_app = EmployeeMaster.objects.filter(employee_email_id=approver_email, deleted_status=False).first()
                        if emp_app:
                            approver_login = emp_app.domain_id
                        if not approver_login:
                            approver_login = approver_email or approver_label
                        
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "UPDATE tbl_UploadSuccessLog SET Approver = %s WHERE LogId = %s",
                                [approver_login, int(request_id)]
                            )
                except Exception as app_err:
                    logging.error(f"Error updating Approver column: {app_err}")

                try:
                    abs_file_path = Path(settings.BASE_DIR) / file_path.lstrip('/')
                    analyze_and_generate_flags(int(request_id), str(abs_file_path), doc_code, user_geo)
                except Exception as flag_err:
                    logging.error(f"Error generating flags during upload: {flag_err}")
                
                try:
                    send_request_submit_email(request, request_id, doc_code, user_geo)
                except Exception as mail_err:
                    logging.error(f"Error sending submit email: {mail_err}")
        return JsonResponse(
            {
                'success': True,
                'source': 'database',
                'message': 'success : Request uploaded successfully',
                'requestId': request_id,
                'filePath': file_path,
                'emailPath': email_path,
            }
        )
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return JsonResponse(
            {
                'success': False,
                'source': 'local-file',
                'message': f'Database Insertion Error: {str(exc)}',
                'filePath': file_path,
                'emailPath': email_path,
            },
            status=500
        )


@csrf_exempt
@require_POST
def validate_upload(request):
    template_file = request.FILES.get('templateFile')
    email_file = request.FILES.get('emailFile')
    doc_code = normalize_document_type(request.POST.get('documentType'))
    user_geo = request.POST.get('usergeo') or get_session_geo(request)

    checklist, overall_success = run_validation_checklist(template_file, email_file, doc_code, user_geo)
    
    return JsonResponse({
        'success': overall_success,
        'checklist': checklist
    })


def parse_action_rows(request):
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}
    rows = payload.get('rows') or []
    action = normalize_document_type(payload.get('documentType') or payload.get('action'))
    comments = safe_text(payload.get('comments'))
    return payload, rows, action, comments


def get_admins_by_geo(user_geo):
    from django.conf import settings
    from login.models import EmployeeMaster
    
    geo = str(user_geo or '').strip().lower()
    
    # Map input geo string to standard geo names
    if geo in ('india', 'ind'):
        std_geo = 'India'
    elif geo in ('philippines', 'php'):
        std_geo = 'Philippines'
    elif geo in ('us', 'united states', 'usa'):
        std_geo = 'United States'
    else:
        std_geo = user_geo or 'India' # Fallback
        
    # First attempt: Query admins with std_geo from database
    admin_emails = list(
        EmployeeMaster.objects.filter(role=2, emp_geo__iexact=std_geo, deleted_status=False)
        .values_list('employee_email_id', flat=True)
    )
    admin_emails = [e for e in admin_emails if e and '@' in str(e)]
    
    # Second attempt: Check fallback settings
    if not admin_emails:
        fallback_map = getattr(settings, 'ADMIN_EMAILS_BY_GEO', {})
        admin_emails = fallback_map.get(std_geo) or fallback_map.get(std_geo.lower()) or []
        
    # Ultimate fallback to make sure email list is not empty
    if not admin_emails:
        admin_emails = ['admin@firstsource.com']
        
    return admin_emails


def get_request_details_for_notifications(log_ids):
    if not log_ids:
        return []
        
    log_ids_str = ", ".join(str(int(lid)) for lid in log_ids)
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        # Try finding in tbl_UploadSuccessLog
        cursor.execute(f"""
            SELECT LogId, UploadedBy, UserGeo, DocumentType 
            FROM tbl_UploadSuccessLog 
            WHERE LogId IN ({log_ids_str})
        """)
        rows = cursor.fetchall()
        if rows:
            return [{'id': r[0], 'uploaded_by': r[1], 'user_geo': r[2], 'doc_type': r[3]} for r in rows]
            
        # Try finding in tbl_UploadSuccessLog_MergeMaster
        cursor.execute(f"""
            SELECT MMLogId, CreatedBy, UserGeo, DocType 
            FROM tbl_UploadSuccessLog_MergeMaster 
            WHERE MMLogId IN ({log_ids_str})
        """)
        rows = cursor.fetchall()
        return [{'id': r[0], 'uploaded_by': r[1], 'user_geo': r[2], 'doc_type': r[3]} for r in rows]
    except Exception as e:
        logging.error(f"Error fetching request details for notifications: {e}")
        return []
    finally:
        conn.close()


def send_request_submit_email(request, request_id, doc_code, user_geo):
    import re
    from django.conf import settings
    from django.core.mail import EmailMessage
    from login.models import EmployeeMaster
    
    approver_label = request.POST.get('approver')
    approver_email = ''
    if approver_label:
        match = re.search(r'[\w\.-]+@[\w\.-]+', approver_label)
        if match:
            approver_email = match.group(0).strip().lower()
            
    if not approver_email:
        # Fallback to supervisor or manager email of requestor if approver email not found
        domain_id = request.session.get('DomainId') or request.session.get('LoginID')
        if domain_id:
            emp = EmployeeMaster.objects.filter(domain_id=domain_id, deleted_status=False).first()
            if emp:
                approver_email = emp.supervisor_email_id or emp.manager_email_id or ''
                
    if not approver_email:
        approver_email = 'approver@firstsource.com' # Ultimate fallback
        
    cc_emails = get_admins_by_geo(user_geo)
    
    # Read template
    template_path = os.path.join(settings.BASE_DIR, 'Content', 'Mail', 'EmailNotificationPrimary.txt')
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            text = f.read()
    else:
        text = "Hi, Your Request ID [@@REQUESTID] for Document Type - [@@DOCTYPE] has been processed."
        
    # Replaces
    doc_folder = DOCUMENT_FOLDERS.get(doc_code, {}).get(user_geo, DOCUMENT_FOLDERS.get(doc_code, {}).get('India', ''))
    text = text.replace('[@@DOCTYPE]', doc_folder)
    text = text.replace('[@@REQUESTID]', str(request_id))
    
    subject = "FirstPay Notification - Submit"
    
    msg = EmailMessage(
        subject=subject,
        body=text,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[approver_email],
        cc=cc_emails,
    )
    msg.content_subtype = 'html'
    msg.send(fail_silently=True)


def send_approver_email_notification(request, action, document_type, rows, comments=''):
    log_ids = [int(row['id']) for row in rows]
    details = get_request_details_for_notifications(log_ids)
    
    from django.conf import settings
    from django.core.mail import EmailMessage
    from login.models import EmployeeMaster
    
    # Map action to template file
    if action == 'approve':
        template_name = 'EmailNotificationPrimaryApprove.txt'
        default_subject = "FirstPay Notification - Approve"
    elif action == 'reject':
        template_name = 'EmailNotificationPrimaryReject.txt'
        default_subject = "FirstPay Notification - Reject"
    elif action == 'consolidate':
        template_name = 'EmailNotificationPrimaryConsolidate.txt'
        default_subject = "FirstPay Notification - Consolidate"
    else:
        template_name = 'EmailNotificationPrimary.txt'
        default_subject = "FirstPay Notification - Submit"
        
    template_path = os.path.join(settings.BASE_DIR, 'Content', 'Mail', template_name)
    
    # Group details by requestor to avoid sending multiple separate emails
    from collections import defaultdict
    grouped = defaultdict(list)
    for detail in details:
        key = (detail['uploaded_by'], detail['user_geo'], detail['doc_type'])
        grouped[key].append(detail['id'])
        
    for (uploaded_by, user_geo, doc_type), req_ids in grouped.items():
        requestor_email = ''
        if uploaded_by:
            emp = EmployeeMaster.objects.filter(domain_id=uploaded_by, deleted_status=False).first()
            if emp:
                requestor_email = emp.employee_email_id
        if not requestor_email:
            requestor_email = 'requestor@firstsource.com'
            
        cc_emails = get_admins_by_geo(user_geo)
        
        # Read template text
        if os.path.exists(template_path):
            with open(template_path, 'r', encoding='utf-8') as f:
                text = f.read()
        else:
            text = f"Hi, request {req_ids} has been {action}d."
            
        # Replaces
        doc_folder = DOCUMENT_FOLDERS.get(doc_type, {}).get(user_geo, DOCUMENT_FOLDERS.get(doc_type, {}).get('India', ''))
        text = text.replace('[@@DOCTYPE]', doc_folder)
        
        req_ids_str = ", ".join(str(rid) for rid in req_ids)
        text = text.replace('[@@REQUESTID]', req_ids_str)
        
        if action == 'reject':
            text = text.replace('[@@COMMENTS]', comments or 'No comments provided.')
            
        subject = default_subject
        
        msg = EmailMessage(
            subject=subject,
            body=text,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[requestor_email],
            cc=cc_emails,
        )
        msg.content_subtype = 'html'
        msg.send(fail_silently=True)



def merge_excel_files(file_paths, doc_code, output_path):
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    if doc_code in ('IRF', 'TDCT'):
        # Create a new workbook and single sheet
        merged_wb = openpyxl.Workbook()
        merged_ws = merged_wb.active
        merged_ws.title = "MergedSheet"
        
        current_row = 1
        is_first_file = True
        
        for fp in file_paths:
            # Resolve absolute path on disk
            rel_path = fp.replace('\\', '/').lstrip('/')
            abs_path = Path(settings.BASE_DIR) / rel_path
            if not abs_path.exists():
                abs_path = Path(fp)
                if not abs_path.exists():
                    logging.warning(f"Excel merge: file path not found: {fp}")
                    continue
            
            try:
                wb = openpyxl.load_workbook(str(abs_path), data_only=True)
                ws = wb.worksheets[0]
                
                start_row = 1
                if not is_first_file:
                    start_row = 2 # skip header
                    
                for r in range(start_row, ws.max_row + 1):
                    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                    if all(v is None or str(v).strip() == '' for v in row_vals):
                        continue
                        
                    for c_idx, val in enumerate(row_vals, 1):
                        merged_ws.cell(row=current_row, column=c_idx, value=val)
                    current_row += 1
                is_first_file = False
            except Exception as e:
                logging.error(f"Error merging file {fp}: {e}")
                raise e
                
        merged_wb.save(str(output_path))
        
    elif doc_code == 'PRL':
        merged_wb = openpyxl.Workbook()
        default_sheet = merged_wb.active
        merged_wb.remove(default_sheet)
        
        sheets_created = False
        merged_sheets = {}
        
        for fp in file_paths:
            rel_path = fp.replace('\\', '/').lstrip('/')
            abs_path = Path(settings.BASE_DIR) / rel_path
            if not abs_path.exists():
                abs_path = Path(fp)
                if not abs_path.exists():
                    logging.warning(f"Excel merge: PRL path not found: {fp}")
                    continue
            
            try:
                wb = openpyxl.load_workbook(str(abs_path), data_only=True)
                if not sheets_created:
                    for sheet in wb.worksheets:
                        sheet_name = sheet.title.strip()
                        new_ws = merged_wb.create_sheet(title=sheet_name)
                        merged_sheets[sheet_name] = new_ws
                    sheets_created = True
                    
                for sheet in wb.worksheets:
                    sheet_name = sheet.title.strip()
                    merged_ws = merged_sheets.get(sheet_name)
                    if not merged_ws:
                        continue
                        
                    is_first_file = (fp == file_paths[0])
                    start_row = 1
                    if not is_first_file:
                        start_row = 2 # skip header
                        
                    current_merged_row = merged_ws.max_row
                    if current_merged_row == 1 and merged_ws.cell(row=1, column=1).value is None:
                        current_merged_row = 1
                    else:
                        current_merged_row += 1
                        
                    for r in range(start_row, sheet.max_row + 1):
                        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                        if all(v is None or str(v).strip() == '' for v in row_vals):
                            continue
                            
                        for c_idx, val in enumerate(row_vals, 1):
                            merged_ws.cell(row=current_merged_row, column=c_idx, value=val)
                        current_merged_row += 1
            except Exception as e:
                logging.error(f"Error merging PRL file {fp}: {e}")
                raise e
                
        merged_wb.save(str(output_path))


def execute_action_db(proc_name, log_ids, action_name, role_id, domain_id, comments=None):
    if not log_ids:
        return 0
        
    log_ids_str = ", ".join(f"({int(lid)})" for lid in log_ids)
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        if proc_name == 'SP_RejectRecords':
            sql = f"""
            DECLARE @TVP dbo.ListLogIds;
            INSERT INTO @TVP (LogIds) VALUES {log_ids_str};
            EXEC SP_RejectRecords @TVP, ?, ?, ?, ?;
            """
            cursor.execute(sql, [action_name, str(role_id), domain_id, comments or ''])
        elif proc_name == 'SP_ApproveRecords':
            sql = f"""
            DECLARE @TVP dbo.ListLogIds;
            INSERT INTO @TVP (LogIds) VALUES {log_ids_str};
            EXEC SP_ApproveRecords @TVP, ?, ?, ?;
            """
            cursor.execute(sql, [action_name, str(role_id), domain_id])
        elif proc_name == 'SP_ConsolidatedRecords':
            sql = f"""
            DECLARE @TVP dbo.ListLogIds;
            INSERT INTO @TVP (LogIds) VALUES {log_ids_str};
            EXEC SP_ConsolidatedRecords @TVP, ?, ?, ?;
            """
            cursor.execute(sql, [action_name, str(role_id), domain_id])
            
        conn.commit()
        return len(log_ids)
    finally:
        conn.close()


def execute_insert_merger_log(doc_type, merge_file_path, file_name, upload_by, action, role, emp_no, log_ids, remarks):
    if not log_ids:
        return 0
        
    log_ids_str = ", ".join(f"({int(lid)})" for lid in log_ids)
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        sql = f"""
        DECLARE @TVP dbo.ListLogIds;
        INSERT INTO @TVP (LogIds) VALUES {log_ids_str};
        EXEC SP_InsertMergerLog ?, ?, ?, ?, ?, ?, ?, @TVP, ?;
        """
        cursor.execute(sql, [
            doc_type,
            merge_file_path,
            file_name,
            upload_by,
            action,
            str(role),
            emp_no,
            remarks
        ])
        conn.commit()
        return len(log_ids)
    finally:
        conn.close()


def execute_approve_records_db(log_ids, role_id, domain_id, doc_code):
    if not log_ids:
        return 0
    log_ids_str = ", ".join(str(int(lid)) for lid in log_ids)
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        # Determine if individual logs or merge master logs
        cursor.execute(f"SELECT LogId FROM tbl_UploadSuccessLog WHERE LogId IN ({log_ids_str})")
        found_individual_ids = [r[0] for r in cursor.fetchall()]
        
        if found_individual_ids:
            # Individual logs!
            new_state_id = 10 if str(role_id) == '3' else 5
            remarks = 'Approved by Approver' if str(role_id) == '3' else 'Approved by Admin'
            if str(role_id) == '3':
                cursor.execute(f"""
                    UPDATE tbl_UploadSuccessLog 
                    SET StateId = ?, ModifiedDate = GETDATE(), ModifiedBy = ?,
                        ApprovedRejectDate = GETDATE()
                    WHERE IsDeleted = 0 AND LogId IN ({log_ids_str})
                """, [new_state_id, domain_id])
            else:
                cursor.execute(f"""
                    UPDATE tbl_UploadSuccessLog 
                    SET StateId = ?, ModifiedDate = GETDATE(), ModifiedBy = ?,
                        AdminApproverRejectBy = ?, AdminApproverejectDate = GETDATE()
                    WHERE IsDeleted = 0 AND LogId IN ({log_ids_str})
                """, [new_state_id, domain_id, domain_id])
                
            cursor.execute("""
                INSERT INTO tbl_RequestLogger (RequestRole, RequestDomainId, RequestDocType, RequestRemarks, RequestCreatedDate)
                VALUES (?, ?, ?, ?, GETDATE())
            """, [str(role_id), domain_id, doc_code, remarks])
        else:
            # Merge master logs!
            cursor.execute(f"""
                UPDATE tbl_UploadSuccessLog_MergeMaster 
                SET StateId = 5, ModifiedDate = GETDATE(), ModifiedBy = ?
                WHERE IsDeleted = 0 AND MMLogId IN ({log_ids_str})
            """, [domain_id])
            cursor.execute(f"""
                UPDATE tbl_UploadSuccessLog 
                SET StateId = 5, ModifiedDate = GETDATE(), ModifiedBy = ?,
                    AdminApproverRejectBy = ?, AdminApproverejectDate = GETDATE()
                WHERE IsDeleted = 0 AND mergerId IN ({log_ids_str})
            """, [domain_id, domain_id])
            cursor.execute("""
                INSERT INTO tbl_RequestLogger (RequestRole, RequestDomainId, RequestDocType, RequestRemarks, RequestCreatedDate)
                VALUES (?, ?, ?, 'Merger Approve by Admin', GETDATE())
            """, [str(role_id), domain_id, doc_code])
        conn.commit()
        return len(log_ids)
    finally:
        conn.close()


def execute_reject_records_db(log_ids, role_id, domain_id, doc_code, comments):
    if not log_ids:
        return 0
    log_ids_str = ", ".join(str(int(lid)) for lid in log_ids)
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        # Determine if individual logs or merge master logs
        cursor.execute(f"SELECT LogId FROM tbl_UploadSuccessLog WHERE LogId IN ({log_ids_str})")
        found_individual_ids = [r[0] for r in cursor.fetchall()]
        
        if found_individual_ids:
            # Individual logs!
            new_state_id = 11 if str(role_id) == '3' else 4
            remarks = 'Reject by Approver' if str(role_id) == '3' else 'Reject by Admin'
            if str(role_id) == '3':
                cursor.execute(f"""
                    UPDATE tbl_UploadSuccessLog 
                    SET StateId = ?, ModifiedDate = GETDATE(), ModifiedBy = ?, Remarks = ?,
                        ApprovedRejectDate = GETDATE()
                    WHERE IsDeleted = 0 AND LogId IN ({log_ids_str})
                """, [new_state_id, domain_id, comments or ''])
            else:
                cursor.execute(f"""
                    UPDATE tbl_UploadSuccessLog 
                    SET StateId = ?, ModifiedDate = GETDATE(), ModifiedBy = ?, Remarks = ?,
                        AdminApproverRejectBy = ?, AdminApproverejectDate = GETDATE()
                    WHERE IsDeleted = 0 AND LogId IN ({log_ids_str})
                """, [new_state_id, domain_id, comments or '', domain_id])
                
            cursor.execute("""
                INSERT INTO tbl_RequestLogger (RequestRole, RequestDomainId, RequestDocType, RequestRemarks, RequestCreatedDate)
                VALUES (?, ?, ?, ?, GETDATE())
            """, [str(role_id), domain_id, doc_code, remarks])
        else:
            # Merge master logs!
            cursor.execute(f"""
                UPDATE tbl_UploadSuccessLog_MergeMaster 
                SET StateId = 4, ModifiedDate = GETDATE(), ModifiedBy = ?
                WHERE IsDeleted = 0 AND MMLogId IN ({log_ids_str})
            """, [domain_id])
            cursor.execute(f"""
                UPDATE tbl_UploadSuccessLog 
                SET StateId = 4, ModifiedDate = GETDATE(), ModifiedBy = ?, Remarks = ?,
                    AdminApproverRejectBy = ?, AdminApproverejectDate = GETDATE()
                WHERE IsDeleted = 0 AND mergerId IN ({log_ids_str})
            """, [domain_id, comments or '', domain_id])
            cursor.execute("""
                INSERT INTO tbl_RequestLogger (RequestRole, RequestDomainId, RequestDocType, RequestRemarks, RequestCreatedDate)
                VALUES (?, ?, ?, 'Merger Reject by Admin', GETDATE())
            """, [str(role_id), domain_id, doc_code])
        conn.commit()
        return len(log_ids)
    finally:
        conn.close()


@csrf_exempt
@require_POST
def admin_approve(request):
    _, rows, action_name, comments = parse_action_rows(request)
    if not rows:
        return JsonResponse({'success': False, 'message': 'No records selected.'}, status=400)
        
    log_ids = [int(row['id']) for row in rows]
    role_id = request.session.get('RoleID') or '2'
    domain_id = request.session.get('DomainId') or 'system'
    doc_code = DOCUMENT_TYPE_MAP.get(action_name.lower()) or action_name.upper()
    
    try:
        execute_approve_records_db(
            log_ids,
            role_id,
            domain_id,
            doc_code
        )
        send_approver_email_notification(request, 'approve', doc_code, rows, comments)
        return JsonResponse({'success': True, 'message': 'Approved', 'documentType': action_name, 'affected': len(rows)})
    except Exception as exc:
        logging.error(f"Database error during approval: {exc}")
        return JsonResponse({'success': False, 'message': f'Database error during approval: {str(exc)}'}, status=500)


@csrf_exempt
@require_POST
def admin_reject(request):
    _, rows, action_name, comments = parse_action_rows(request)
    if not rows:
        return JsonResponse({'success': False, 'message': 'No records selected.'}, status=400)
    if not comments:
        return JsonResponse({'success': False, 'message': 'Comments are required for rejection.'}, status=400)
        
    log_ids = [int(row['id']) for row in rows]
    role_id = request.session.get('RoleID') or '2'
    domain_id = request.session.get('DomainId') or 'system'
    doc_code = DOCUMENT_TYPE_MAP.get(action_name.lower()) or action_name.upper()
    
    try:
        execute_reject_records_db(
            log_ids,
            role_id,
            domain_id,
            doc_code,
            comments
        )
        send_approver_email_notification(request, 'reject', doc_code, rows, comments)
        return JsonResponse({'success': True, 'message': 'Rejected', 'documentType': action_name, 'affected': len(rows)})
    except Exception as exc:
        logging.error(f"Database error during rejection: {exc}")
        return JsonResponse({'success': False, 'message': f'Database error during rejection: {str(exc)}'}, status=500)


@csrf_exempt
@require_POST
def admin_consolidate(request):
    _, rows, action_name, comments = parse_action_rows(request)
    if not rows:
        return JsonResponse({'success': False, 'message': 'No records selected for consolidation.'}, status=400)
        
    log_ids = [int(row['id']) for row in rows]
    file_paths = [str(row['filePath']) for row in rows]
    
    doc_code = DOCUMENT_TYPE_MAP.get(action_name.lower()) or action_name.upper()
    role_id = request.session.get('RoleID') or '2'
    domain_id = request.session.get('DomainId') or 'system'
    login_id = request.session.get('LoginID') or 'system'
    user_geo = get_session_geo(request)
    
    timestamp_str = timezone.now().strftime("%d%m%Y%H%M%S")
    file_name = f"{doc_code}{domain_id}_{timestamp_str}Merge.xlsx"
    
    merge_folder = 'PhpMerge' if user_geo == 'Philippines' else 'IndiaMerge'
    relative_db_path = f"/Content/MergedFiles/{merge_folder}/{file_name}"
    
    output_dir = settings.BASE_DIR / 'Content' / 'MergedFiles' / merge_folder
    output_filepath = output_dir / file_name
    
    try:
        merge_excel_files(file_paths, doc_code, output_filepath)
    except Exception as e:
        logging.error(f"Error merging files: {e}")
        return JsonResponse({'success': False, 'message': f'Failed to merge Excel files: {str(e)}'}, status=500)
        
    remarks_str = ",".join(str(lid) for lid in log_ids)
    
    try:
        execute_insert_merger_log(
            doc_code,
            relative_db_path,
            file_name,
            domain_id,
            'Insert',
            role_id,
            login_id,
            log_ids,
            remarks_str
        )
        
        execute_action_db(
            'SP_ConsolidatedRecords',
            log_ids,
            doc_code,
            role_id,
            domain_id
        )
        
        try:
            send_approver_email_notification(request, 'consolidate', doc_code, rows, comments)
        except Exception as mail_err:
            logging.error(f"Error sending consolidation email: {mail_err}")
            
        return JsonResponse({
            'success': True,
            'message': 'Consolidated successfully!',
            'documentType': action_name,
            'affected': len(log_ids)
        })
    except Exception as exc:
        logging.error(f"Database error during consolidation: {exc}")
        return JsonResponse({
            'success': False,
            'message': f'Database error during consolidation: {str(exc)}'
        }, status=500)


@require_GET
def db_status(request):
    try:
        database_status = check_database_connection()
        return JsonResponse({'success': True, **database_status})
    except DatabaseError as exc:
        return JsonResponse(
            {
                'success': False,
                'connected': False,
                'message': str(exc),
                'configuredServer': settings.PAYROLL_SQL_SERVER,
                'configuredDatabase': settings.PAYROLL_SQL_DATABASE,
            },
            status=503,
        )


@csrf_exempt
@require_POST
def domain_login(request):
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid request payload.'}, status=400)

    domain_name = str(payload.get('domainName', '')).strip()
    domain_id = str(payload.get('domainId', '')).strip()
    password = str(payload.get('password', ''))

    errors = []
    if not domain_name or domain_name == '0':
        errors.append('Please select domain name.')
    if not domain_id:
        errors.append("Domain id can't be blank")
    if not password:
        errors.append("Password can't be blank")

    if errors:
        return JsonResponse({'success': False, 'message': ' '.join(errors)}, status=400)

    try:
        auth_status = auth_service_method(domain_name, domain_id, password)

        if auth_status.upper() == 'FALSE' and is_auth_bypass_user(domain_id):
            auth_status = 'TRUE'

        if auth_status.upper() != 'TRUE' and auth_status != 'abort':
            return login_failure()

        employee = get_employee_by_domain_id(domain_id)
    except DatabaseError:
        return JsonResponse(
            {
                'success': False,
                'message': (
                    f'Unable to connect to {settings.PAYROLL_SQL_DATABASE} '
                    f'on {settings.PAYROLL_SQL_SERVER}.'
                ),
            },
            status=503,
        )
    except Exception:
        return login_failure(status=500)

    if employee is None:
        return login_failure(LOGIN_FAILURE_MESSAGE, status=403)

    store_login_session(request, employee)
    user_payload = build_user_payload(domain_name, employee)

    return JsonResponse(
        {
            'success': True,
            'message': 'Login successful.',
            'user': user_payload,
        }
    )


@csrf_exempt
@require_POST
def sign_out(request):
    request.session.flush()
    return JsonResponse({'success': True, 'message': 'Signed out successfully.'})


@require_GET
def session_info(request):
    if not request.session.get('LoginID'):
        return JsonResponse({'authenticated': False}, status=401)

    return JsonResponse(
        {
            'authenticated': True,
            'user': {
                'employeeCode': request.session.get('LoginID'),
                'employeeName': request.session.get('EmployeeName'),
                'supervisor': request.session.get('Supervisor'),
                'role': request.session.get('RoleID'),
                'grade': request.session.get('Grade'),
                'domainId': request.session.get('DomainId'),
                'emailId': request.session.get('EmailId'),
                'empGeo': request.session.get('EmpGeo'),
                'loginType': request.session.get('LoginType'),
                'rootDir': request.session.get('RootDir'),
            },
        }
    )


# =========================================================================
# SPOC MANUAL VALIDATION ENDPOINTS (PHASE 3)
# =========================================================================

def remove_declined_rows_from_excel(log_id):
    """
    Finds accepted flags (declined rows) for the given log ID,
    loads the Excel file, deletes those rows, and saves the file.
    """
    from .db import get_payroll_connection
    import openpyxl
    
    conn = get_payroll_connection()
    file_path = None
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT FilePath FROM tbl_UploadSuccessLog WHERE LogId = ?", [log_id])
        row = cursor.fetchone()
        if row:
            file_path = row[0]
    except Exception as e:
        logging.error(f"Error fetching FilePath: {e}")
        
    if not file_path:
        conn.close()
        return

    accepted_flags = []
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT SheetName, RowIndex, EmpNo, PayoutMonth, Amount 
            FROM tbl_UploadRowFlags 
            WHERE LogId = ? AND SpocAction = 'Accepted'
        """, [log_id])
        accepted_flags = cursor.fetchall()
    except Exception as e:
        logging.error(f"Error fetching accepted flags: {e}")
    finally:
        conn.close()

    if not accepted_flags:
        return

    rel_path = file_path.replace('\\', '/').lstrip('/')
    abs_path = os.path.join(settings.BASE_DIR, rel_path)
    if not os.path.exists(abs_path):
        return

    try:
        wb = openpyxl.load_workbook(abs_path)
        
        from collections import defaultdict
        sheet_rows = defaultdict(list)
        for sheet_name, row_idx, emp_no, payout_month, amount in accepted_flags:
            sheet_rows[sheet_name].append(row_idx)

        for sheet_name, row_indices in sheet_rows.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for r_idx in sorted(row_indices, reverse=True):
                    if r_idx <= ws.max_row:
                        ws.delete_rows(r_idx, 1)

        wb.save(abs_path)
    except Exception as e:
        logging.error(f"Error removing declined rows from excel: {e}")


@api_view(['GET'])
def get_upload_flags(request, log_id):
    """
    Fetches the row-level validation flags generated for a specific upload request.
    """
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT FlagId, SheetName, RowIndex, EmpNo, EmpName, PayoutType, PayoutMonth, Amount, BankAccount, FlagType, FlagMessage, SpocAction, SpocComment
            FROM tbl_UploadRowFlags 
            WHERE LogId = ?
            ORDER BY SheetName, RowIndex
        """, [log_id])
        rows = cursor.fetchall()
        flags = []
        for r in rows:
            flags.append({
                "flagId": r[0],
                "sheetName": r[1],
                "rowIndex": r[2],
                "empNo": r[3],
                "empName": r[4],
                "payoutType": r[5],
                "payoutMonth": r[6],
                "amount": float(r[7]) if r[7] is not None else 0.0,
                "bankAccount": r[8],
                "flagType": r[9],
                "flagMessage": r[10],
                "spocAction": r[11],
                "spocComment": r[12]
            })
        return JsonResponse({"success": True, "flags": flags})
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
    finally:
        conn.close()


@csrf_exempt
@api_view(['POST'])
def submit_flag_decisions(request, log_id):
    """
    Receives SPOC decisions for the validation flags of an upload request.
    If a flag is 'Accepted', the row is declined and deleted from the saved Excel file.
    """
    login_id = get_session_login_id(request)
    payload = request.data or {}
    decisions = payload.get('decisions') or []

    if not decisions:
        return Response({"success": False, "message": "No decisions provided."}, status=400)

    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        for dec in decisions:
            flag_id = dec.get('flagId')
            action = dec.get('action')  # 'Accepted' (Row Declined) or 'Declined' (Allowed Exception)
            comment = dec.get('comment', '')

            cursor.execute("""
                UPDATE tbl_UploadRowFlags
                SET SpocAction = ?, SpocComment = ?, ActionedBy = ?, ActionedDate = GETDATE()
                WHERE FlagId = ? AND LogId = ?
            """, [action, comment, login_id, flag_id, log_id])

        conn.commit()
    except Exception as e:
        return Response({"success": False, "message": f"Database error: {str(e)}"}, status=500)
    finally:
        conn.close()

    # Rebuild Excel file by removing declined rows
    try:
        remove_declined_rows_from_excel(log_id)
    except Exception as e:
        logging.error(f"Failed to rebuild Excel file after SPOC flag action: {e}")

    # Count accepted flags (declined rows) and update success log remarks
    try:
        conn = get_payroll_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM tbl_UploadRowFlags WHERE LogId = ? AND SpocAction = 'Accepted'", [log_id])
        declined_count = cursor.fetchone()[0]
        
        if declined_count > 0:
            remarks = f"Completed with {declined_count} declined rows."
            cursor.execute("""
                UPDATE tbl_UploadSuccessLog
                SET Remarks = ?
                WHERE LogId = ?
            """, [remarks, log_id])
            conn.commit()
    except Exception as e:
        pass
    finally:
        conn.close()

    return Response({"success": True, "message": "Decisions saved successfully."})


@api_view(['GET'])
def get_declined_rows_summary(request, log_id):
    """
    Returns a summary of rows declined (flag accepted) by the SPOC for display on the OPS dashboard.
    """
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT FlagId, SheetName, RowIndex, EmpNo, EmpName, PayoutType, PayoutMonth, Amount, FlagMessage, SpocComment
            FROM tbl_UploadRowFlags
            WHERE LogId = ? AND SpocAction = 'Accepted'
            ORDER BY SheetName, RowIndex
        """, [log_id])
        rows = cursor.fetchall()
        declined_rows = []
        for r in rows:
            declined_rows.append({
                "flagId": r[0],
                "sheetName": r[1],
                "rowIndex": r[2],
                "empNo": r[3],
                "empName": r[4],
                "payoutType": r[5],
                "payoutMonth": r[6],
                "amount": float(r[7]) if r[7] is not None else 0.0,
                "reason": r[8],
                "spocComment": r[9]
            })
        return JsonResponse({"success": True, "declinedRows": declined_rows})
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
    finally:
        conn.close()

