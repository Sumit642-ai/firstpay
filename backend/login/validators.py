import os
import re
from datetime import datetime, date
import openpyxl
from django.core.exceptions import ValidationError

EXPECTED_TEMPLATES = {
    'PRL': {
        'India': {
            'sheets_count': 10,
            'sheets_names': [
                'VP Ratings', 'VP Arrears', 'Incentives', 'Incentive Arrears', 
                'Split shift', 'Other Allowances', 'Other Allowance - Arrears', 
                'Stop Pay', 'Resignees', 'Firstsource Confidential'
            ],
            'columns': {
                'VP Ratings': ['Emp No', 'Employee Name', 'DOJ', 'Grade', 'Process Name', 'Score', 'Ratings Month', 'Ratings', 'Amounts', 'Remarks', 'Payment/Deduction'],
                'VP Arrears': ['Emp No', 'Employee Name', 'DOJ', 'Grade', 'Process Name', 'Arrears Month', 'Score', 'Ratings', 'Amounts', 'Remarks', 'Payment/Deduction'],
                'Incentives': ['Emp No', 'Employee Name', 'Process', 'Incentive Month', 'Type of Incentives', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Incentive Arrears': ['Emp No', 'Employee Name', 'Process', 'Incentive Month', 'Type of Incentives', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Split shift': ['Emp No', 'Employee Name', 'Process Name', 'No of Split Shifts', 'Month', 'Split Shift Amount', 'Split Shift Smount Arrears', 'Remarks'],
                'Other Allowances': ['Emp No', 'Employee Name', 'Process', 'Allowances Month', 'Type of Allowances', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Other Allowance - Arrears': ['Emp No', 'Employee Name', 'Process', 'Allowances Month', 'Type of Allowances', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Stop Pay': ['Emp No', 'Employee Name', 'Salary Status', 'Reason for Stop Pay'],
                'Resignees': ['Emp No', 'Employee Name', 'LWD', 'Remarks'],
                'Firstsource Confidential': []
            },
            'allowed_empty': {
                'VP Ratings': ['DOJ', 'Grade', 'Process Name', 'Remarks'],
                'VP Arrears': ['DOJ', 'Grade', 'Process Name', 'Remarks'],
                'Incentives': ['Process', 'Remarks'],
                'Incentive Arrears': ['Process', 'Remarks'],
                'Split shift': ['Process Name', 'Split Shift Smount Arrears', 'Remarks'],
                'Other Allowances': ['Process', 'Remarks'],
                'Other Allowance - Arrears': ['Process', 'Remarks'],
                'Stop Pay': [],
                'Resignees': ['Remarks'],
                'Firstsource Confidential': []
            }
        },
        'Philippines': {
            'sheets_count': 9,
            'sheets_names': [
                'Variablepay Incentive', 'Variablepay Incentive - Arrears', 'Incentives', 
                'Incentives - Arrears', 'Internet allowance & arrears', 'Transport Allowance - Arrears', 
                'Salary Hold', 'Reactivation', 'Firstsource Confidential'
            ],
            'columns': {
                'Variablepay Incentive': ['Emp No', 'Employee Name', 'DOJ', 'Grade', 'Process Name', 'Score', 'Ratings Month', 'Ratings', 'Amounts', 'Remarks', 'Payment/Deduction'],
                'Variablepay Incentive - Arrears': ['Emp No', 'Employee Name', 'DOJ', 'Grade', 'Process Name', 'Arrears Month', 'Score', 'Ratings', 'Amounts', 'Remarks', 'Payment/Deduction'],
                'Incentives': ['Emp No', 'Employee Name', 'Process', 'Incentive Month', 'Type of Incentives', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Incentives - Arrears': ['Emp No', 'Employee Name', 'Process', 'Incentive Month', 'Type of Incentives', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Internet allowance & arrears': ['Emp No', 'Employee Name', 'Process', 'Allowances Month', 'No of worked days', 'Amount', 'Payment/Deduction', 'Remarks'],
                'Transport Allowance - Arrears': ['Emp No', 'Employee Name', 'Process', 'Allowances Month', 'No of worked days', 'Payment/Deduction', 'Remarks'],
                'Salary Hold': ['Emp No', 'Employee Name', 'Date Reported / Requested', 'Last working day', 'Reason(Resignation,RTWO,Termination,Supension,Floating)', 'Remarks'],
                'Reactivation': ['Emp No', 'Employee Name', 'Date Reported / Requested', 'Date Reported to Work', 'Last Working Day', 'Reason(Resignation,RTWO,Termination,Supension,Floating)', 'Remarks'],
                'Firstsource Confidential': []
            },
            'allowed_empty': {
                'Variablepay Incentive': ['DOJ', 'Grade', 'Process Name', 'Remarks'],
                'Variablepay Incentive - Arrears': ['DOJ', 'Grade', 'Process Name', 'Remarks'],
                'Incentives': ['Process', 'Remarks'],
                'Incentives - Arrears': ['Process', 'Remarks'],
                'Internet allowance & arrears': ['Process', 'No of worked days', 'Remarks'],
                'Transport Allowance - Arrears': ['Process', 'No of worked days', 'Remarks'],
                'Salary Hold': ['Date Reported / Requested', 'Remarks'],
                'Reactivation': ['Date Reported / Requested', 'Last Working Day', 'Remarks'],
                'Firstsource Confidential': []
            }
        }
    },
    'IRF': {
        'sheets_count': 2,
        'sheets_names': ['I_Refer', 'Firstsource Confidential'],
        'columns': {
            'I_Refer': ['Employee Name', 'Emp No', 'Designation', 'Grade', 'Process', 'Location', 'Cost Center Code', 'Amount to Pay', 'Name of Joinee', 'Candidate Id', 'Employee Id', 'Designation', 'Grade', 'Estate', 'Process', 'Location', 'Center', 'Cost Center Code', 'Date of Joining', 'Remarks', 'Reason'],
            'Firstsource Confidential': []
        },
        'allowed_empty': {
            'I_Refer': ['Designation', 'Grade', 'Process', 'Location', 'Cost Center Code', 'Estate', 'Center', 'Date of Joining', 'Remarks', 'Reason'],
            'Firstsource Confidential': []
        }
    },
    'TDCT': {
        'sheets_count': 2,
        'sheets_names': ['TransportDeduct', 'Firstsource Confidential'],
        'columns': {
            'TransportDeduct': ['Spoc Name', 'Emp No', 'Employee Name', 'Location', 'CoPay Charges', 'Default Charges', 'Reversal Amount', 'Total Deduction (copay+default-reversal)', 'CoPay Dates considered', 'Default Dates considered', 'Reversal Dates considered'],
            'Firstsource Confidential': []
        },
        'allowed_empty': {
            'TransportDeduct': ['CoPay Charges', 'Default Charges', 'Reversal Amount', 'CoPay Dates considered', 'Default Dates considered', 'Reversal Dates considered'],
            'Firstsource Confidential': []
        }
    }
}

def validate_uploaded_excel(file_obj, doc_code, user_geo):
    """
    Performs comprehensive structural and content validations on the uploaded template file.
    Raises django.core.exceptions.ValidationError if any validation check fails.
    """
    # 1. Invalid Excel file format check
    file_obj.seek(0)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception:
        raise ValidationError("Invalid Excel file format. Could not open the file.")

    # 2. Get expected template configuration
    if doc_code == 'PRL':
        geo = 'Philippines' if str(user_geo).strip().lower() == 'philippines' else 'India'
        config = EXPECTED_TEMPLATES['PRL'][geo]
    else:
        if doc_code not in EXPECTED_TEMPLATES:
            raise ValidationError(f"Unsupported document code '{doc_code}'.")
        config = EXPECTED_TEMPLATES[doc_code]

    actual_sheets = wb.sheetnames
    expected_sheets = config['sheets_names']
    expected_count = config['sheets_count']

    # 3. Check exact number of sheets
    if len(actual_sheets) != expected_count:
        raise ValidationError(f"Sheet count mismatch: expected {expected_count} sheets, but found {len(actual_sheets)}.")

    # 4. Check exact sheet names and order
    for idx, (act_name, exp_name) in enumerate(zip(actual_sheets, expected_sheets)):
        if act_name != exp_name:
            raise ValidationError(f"Sheet name or order mismatch: expected '{exp_name}' at sheet {idx + 1}, but found '{act_name}'.")

    # 5. Check if sheet has formatting or hidden data (hidden worksheets, rows, columns, or protection)
    for sheet_name in actual_sheets:
        ws = wb[sheet_name]
        
        # Check hidden worksheets
        if ws.sheet_state != 'visible':
            raise ValidationError(f"Hidden sheet detected: '{sheet_name}' is hidden or very hidden.")

        # Check sheet protection
        if ws.protection and ws.protection.sheet:
            raise ValidationError(f"Sheet '{sheet_name}' is protected.")

        # Check hidden rows
        for r, dim in ws.row_dimensions.items():
            if dim and dim.hidden:
                raise ValidationError(f"Hidden row detected: Row {r} is hidden in sheet '{sheet_name}'.")

        # Check hidden columns
        for c, dim in ws.column_dimensions.items():
            if dim and dim.hidden:
                raise ValidationError(f"Hidden column detected: Column '{c}' is hidden in sheet '{sheet_name}'.")

    # 6. "Firstsource Confidential" SHEET must exist and be completely blank
    if "Firstsource Confidential" not in actual_sheets:
        raise ValidationError("Missing required sheet: 'Firstsource Confidential'.")
    
    fc_sheet = wb["Firstsource Confidential"]
    for r_idx, row in enumerate(fc_sheet.iter_rows(values_only=True), start=1):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            raise ValidationError("The 'Firstsource Confidential' sheet must be completely blank.")

    # 7. Validation for columns and content rows
    month_regex = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[’']\d{2}$")

    for sheet_name in actual_sheets:
        if sheet_name == "Firstsource Confidential":
            continue

        ws = wb[sheet_name]
        expected_cols = config['columns'][sheet_name]
        allowed_empty = config['allowed_empty'][sheet_name]

        all_rows = list(ws.iter_rows(values_only=True))

        # Find first non-empty row (header row)
        header_row_idx = None
        header_cells = None
        for r_idx, row in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header_row_idx = r_idx
                # Clean headers and strip trailing empty cells
                header_cells = [str(cell).strip() if cell is not None else "" for cell in row]
                while header_cells and header_cells[-1] == "":
                    header_cells.pop()
                break

        if header_row_idx is None:
            raise ValidationError(f"Sheet '{sheet_name}' is empty (no header row found).")

        # Check if there are any empty columns in header
        for c_idx, cell_val in enumerate(header_cells):
            if cell_val == "":
                raise ValidationError(f"Empty column header found at position {c_idx + 1} in sheet '{sheet_name}'.")

        # Column count mismatch
        if len(header_cells) != len(expected_cols):
            raise ValidationError(f"Column count mismatch in sheet '{sheet_name}': expected {len(expected_cols)} columns, but found {len(header_cells)}.")

        # Check exact column names and order
        for c_idx, (act_col, exp_col) in enumerate(zip(header_cells, expected_cols)):
            if act_col != exp_col:
                raise ValidationError(f"Column name or order mismatch in sheet '{sheet_name}' at position {c_idx + 1}: expected '{exp_col}', but found '{act_col}'.")

        # Precompute validation indices for data rows
        required_indices = []
        month_indices = []
        for c_idx, col_name in enumerate(expected_cols):
            if col_name not in allowed_empty:
                required_indices.append(c_idx)
            if "month" in col_name.lower():
                month_indices.append((c_idx, col_name))

        # Validate each data row
        for r_idx in range(header_row_idx + 1, len(all_rows)):
            row = all_rows[r_idx]
            # Skip entirely empty rows
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            # Check valid empty cells for specified listed columns
            for req_idx in required_indices:
                val = row[req_idx] if req_idx < len(row) else None
                if val is None or str(val).strip() == "":
                    raise ValidationError(f"Mandatory column '{expected_cols[req_idx]}' contains an empty cell at row {r_idx + 1} in sheet '{sheet_name}'.")

            # MONTH FORMAT VALIDATION
            for m_idx, col_name in month_indices:
                val = row[m_idx] if m_idx < len(row) else None
                if val is not None and str(val).strip() != "":
                    if isinstance(val, (datetime, date)):
                        val_str = val.strftime("%b'%y")
                    else:
                        val_str = str(val).strip()

                    if not month_regex.match(val_str):
                        raise ValidationError(f"Invalid month format '{val_str}' at row {r_idx + 1}, column '{col_name}' in sheet '{sheet_name}'. Expected format like: Mar’26 or Mar'26.")
    
    # Seek back to 0 just in case view needs to read it
    file_obj.seek(0)

def run_validation_checklist(template_file, email_file, doc_code, user_geo):
    """
    Runs all 12 validations in order and returns a structured checklist
    dict with pass/fail/pending status, and an overall boolean success status.
    """
    checklist = {
        "requiredFields": {"label": "Required Fields Check", "status": "pending", "message": "Checking required fields..."},
        "extensionCheck": {"label": "File Extension Check", "status": "pending", "message": "Checking file extension..."},
        "excelFormat": {"label": "Excel File Format", "status": "pending", "message": "Pending Excel structural check."},
        "sheetCount": {"label": "Sheet Count Check", "status": "pending", "message": "Pending sheet count check."},
        "sheetNames": {"label": "Sheet Names & Order Check", "status": "pending", "message": "Pending sheet names check."},
        "confidentialSheet": {"label": "Confidential Sheet Blank Check", "status": "pending", "message": "Pending blank sheet check."},
        "formattingHidden": {"label": "Formatting & Hidden Data Check", "status": "pending", "message": "Pending formatting check."},
        "columnCount": {"label": "Column Count Check", "status": "pending", "message": "Pending column count check."},
        "columnNames": {"label": "Column Names & Order Check", "status": "pending", "message": "Pending column names check."},
        "headerBlank": {"label": "Empty Columns in Header Check", "status": "pending", "message": "Pending empty header check."},
        "monthFormat": {"label": "Month Format Validation", "status": "pending", "message": "Pending month format check."},
        "emptyCells": {"label": "Mandatory Column Empty Cell Check", "status": "pending", "message": "Pending empty cell check."}
    }
    
    def fail_remaining(message):
        for key in checklist:
            if checklist[key]["status"] == "pending":
                checklist[key]["status"] = "fail"
                checklist[key]["message"] = message

    # 1. Required Fields Check
    if not doc_code or doc_code not in EXPECTED_TEMPLATES:
        checklist["requiredFields"] = {"label": "Required Fields Check", "status": "fail", "message": "select your option !"}
        fail_remaining("Cannot validate due to missing/invalid document type selection.")
        return checklist, False
    
    if template_file is None:
        checklist["requiredFields"] = {"label": "Required Fields Check", "status": "fail", "message": "Template file must be attached."}
        fail_remaining("Cannot validate due to missing template file.")
        return checklist, False
        
    checklist["requiredFields"] = {"label": "Required Fields Check", "status": "pass", "message": "Document type selected and template attached."}

    # 2. Template File Extension Check
    ext = os.path.splitext(template_file.name)[1].lower()
    if ext not in ('.xls', '.xlsx'):
        checklist["extensionCheck"] = {"label": "File Extension Check", "status": "fail", "message": "Only .xls and .xlsx formats allowed."}
        fail_remaining("Cannot validate due to invalid template file extension.")
        return checklist, False
        
    if email_file is not None:
        email_ext = os.path.splitext(email_file.name)[1].lower()
        if email_ext not in ('.msg', '.eml'):
            checklist["extensionCheck"] = {"label": "File Extension Check", "status": "fail", "message": "Only .msg and .eml email formats allowed."}
            fail_remaining("Cannot validate due to invalid email file extension.")
            return checklist, False
        
    checklist["extensionCheck"] = {"label": "File Extension Check", "status": "pass", "message": "Valid file extensions."}

    # 3. Invalid Excel file format check
    template_file.seek(0)
    try:
        wb = openpyxl.load_workbook(template_file, data_only=True)
        checklist["excelFormat"] = {"label": "Excel File Format", "status": "pass", "message": "Excel file is readable."}
    except Exception:
        checklist["excelFormat"] = {"label": "Excel File Format", "status": "fail", "message": "Invalid Excel file format. Could not open the file."}
        fail_remaining("Cannot validate due to unreadable Excel file.")
        return checklist, False

    # 4. Get expected template configuration
    if doc_code == 'PRL':
        geo = 'Philippines' if str(user_geo).strip().lower() == 'philippines' else 'India'
        config = EXPECTED_TEMPLATES['PRL'][geo]
    else:
        config = EXPECTED_TEMPLATES[doc_code]

    actual_sheets = wb.sheetnames
    expected_sheets = config['sheets_names']
    expected_count = config['sheets_count']

    # 5. Check exact number of sheets
    if len(actual_sheets) != expected_count:
        checklist["sheetCount"] = {"label": "Sheet Count Check", "status": "fail", "message": f"Expected {expected_count} sheets, found {len(actual_sheets)}."}
        fail_remaining("Cannot validate sheet contents due to sheet count mismatch.")
        return checklist, False
    checklist["sheetCount"] = {"label": "Sheet Count Check", "status": "pass", "message": f"Sheet count is correct ({expected_count} sheets)."}

    # 6. Check exact sheet names and order
    sheet_mismatch = False
    for idx, (act_name, exp_name) in enumerate(zip(actual_sheets, expected_sheets)):
        if act_name != exp_name:
            checklist["sheetNames"] = {"label": "Sheet Names & Order Check", "status": "fail", "message": f"Expected sheet '{exp_name}' at position {idx+1}, found '{act_name}'."}
            sheet_mismatch = True
            break
    if sheet_mismatch:
        fail_remaining("Cannot validate sheet contents due to sheet name/order mismatch.")
        return checklist, False
    checklist["sheetNames"] = {"label": "Sheet Names & Order Check", "status": "pass", "message": "Sheet names and order match."}

    # 7. Formatting or hidden data check (hidden worksheets, rows, columns, protection)
    hidden_detected = False
    for sheet_name in actual_sheets:
        ws = wb[sheet_name]
        if ws.sheet_state != 'visible':
            checklist["formattingHidden"] = {"label": "Formatting & Hidden Data Check", "status": "fail", "message": f"Sheet '{sheet_name}' is hidden."}
            hidden_detected = True
            break
        if ws.protection and ws.protection.sheet:
            checklist["formattingHidden"] = {"label": "Formatting & Hidden Data Check", "status": "fail", "message": f"Sheet '{sheet_name}' is protected."}
            hidden_detected = True
            break
        for r, dim in ws.row_dimensions.items():
            if dim and dim.hidden:
                checklist["formattingHidden"] = {"label": "Formatting & Hidden Data Check", "status": "fail", "message": f"Row {r} is hidden in sheet '{sheet_name}'."}
                hidden_detected = True
                break
        if hidden_detected:
            break
        for c, dim in ws.column_dimensions.items():
            if dim and dim.hidden:
                checklist["formattingHidden"] = {"label": "Formatting & Hidden Data Check", "status": "fail", "message": f"Column '{c}' is hidden in sheet '{sheet_name}'."}
                hidden_detected = True
                break
        if hidden_detected:
            break
            
    if hidden_detected:
        fail_remaining("Cannot validate contents due to hidden elements or sheet protection.")
        return checklist, False
    checklist["formattingHidden"] = {"label": "Formatting & Hidden Data Check", "status": "pass", "message": "No hidden rows, columns, sheets, or sheet protection detected."}

    # 8. Firstsource Confidential sheet blank check
    if "Firstsource Confidential" not in actual_sheets:
        checklist["confidentialSheet"] = {"label": "Confidential Sheet Blank Check", "status": "fail", "message": "Missing 'Firstsource Confidential' sheet."}
        fail_remaining("Cannot validate due to missing confidential sheet.")
        return checklist, False
    
    fc_sheet = wb["Firstsource Confidential"]
    confidential_blank = True
    
    # Check if sheet has defined range beyond A1 (empty formatting or hidden data not allowed)
    dimension = fc_sheet.calculate_dimension()
    if dimension and dimension not in ("A1:A1", "A1"):
        confidential_blank = False
        
    for row in fc_sheet.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            confidential_blank = False
            break
    if not confidential_blank:
        checklist["confidentialSheet"] = {"label": "Confidential Sheet Blank Check", "status": "fail", "message": "The 'Firstsource Confidential' sheet must be completely blank. Empty formatting or hidden data range beyond A1 is not allowed."}
        fail_remaining("Cannot validate due to data or formatting in confidential sheet.")
        return checklist, False
    checklist["confidentialSheet"] = {"label": "Confidential Sheet Blank Check", "status": "pass", "message": "'Firstsource Confidential' sheet is blank."}

    # Now we loop through sheets to validate columns and rows
    column_count_ok = True
    column_names_ok = True
    header_blank_ok = True
    month_format_ok = True
    empty_cells_ok = True
    
    month_regex = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[’']\d{2}$")

    for sheet_name in actual_sheets:
        if sheet_name == "Firstsource Confidential":
            continue
            
        ws = wb[sheet_name]
        expected_cols = config['columns'][sheet_name]
        allowed_empty = config['allowed_empty'][sheet_name]

        all_rows = list(ws.iter_rows(values_only=True))
        header_row_idx = None
        header_cells = None
        for r_idx, row in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header_row_idx = r_idx
                header_cells = [str(cell).strip() if cell is not None else "" for cell in row]
                while header_cells and header_cells[-1] == "":
                    header_cells.pop()
                break

        if header_row_idx is None:
            checklist["columnCount"] = {"label": "Column Count Check", "status": "fail", "message": f"Sheet '{sheet_name}' is empty."}
            column_count_ok = False
            break

        # Check empty header cells
        for c_idx, cell_val in enumerate(header_cells):
            if cell_val == "":
                checklist["headerBlank"] = {"label": "Empty Columns in Header Check", "status": "fail", "message": f"Empty header column found at column {c_idx+1} in sheet '{sheet_name}'."}
                header_blank_ok = False
                break
        if not header_blank_ok:
            break

        # Check column count
        if len(header_cells) != len(expected_cols):
            checklist["columnCount"] = {"label": "Column Count Check", "status": "fail", "message": f"Column count mismatch in sheet '{sheet_name}': expected {len(expected_cols)}, found {len(header_cells)}."}
            column_count_ok = False
            break

        # Check column names and order
        for c_idx, (act_col, exp_col) in enumerate(zip(header_cells, expected_cols)):
            is_amount_flex = (act_col.lower() in ('amount', 'amounts') and exp_col.lower() in ('amount', 'amounts'))
            if act_col != exp_col and not is_amount_flex:
                checklist["columnNames"] = {"label": "Column Names & Order Check", "status": "fail", "message": f"Mismatch in sheet '{sheet_name}' column {c_idx+1}: expected '{exp_col}', found '{act_col}'."}
                column_names_ok = False
                break
        if not column_names_ok:
            break

        # Precompute validation indices for data rows
        required_indices = []
        month_indices = []
        for c_idx, col_name in enumerate(expected_cols):
            if col_name not in allowed_empty:
                required_indices.append(c_idx)
            if "month" in col_name.lower():
                month_indices.append((c_idx, col_name))

        # Check data rows
        for r_idx in range(header_row_idx + 1, len(all_rows)):
            row = all_rows[r_idx]
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            # Empty cells in mandatory columns
            for req_idx in required_indices:
                val = row[req_idx] if req_idx < len(row) else None
                if val is None or str(val).strip() == "":
                    checklist["emptyCells"] = {"label": "Mandatory Column Empty Cell Check", "status": "fail", "message": f"Required column '{expected_cols[req_idx]}' is empty at row {r_idx+1} in sheet '{sheet_name}'."}
                    empty_cells_ok = False
                    break
            if not empty_cells_ok:
                break

            # Month formatting
            for m_idx, col_name in month_indices:
                val = row[m_idx] if m_idx < len(row) else None
                if val is not None and str(val).strip() != "":
                    if isinstance(val, (datetime, date)):
                        val_str = val.strftime("%b'%y")
                    else:
                        val_str = str(val).strip()
                    if not month_regex.match(val_str):
                        checklist["monthFormat"] = {"label": "Month Format Validation", "status": "fail", "message": f"Invalid month '{val_str}' at row {r_idx+1}, column '{col_name}' in sheet '{sheet_name}'."}
                        month_format_ok = False
                        break
            if not month_format_ok:
                break

        if not empty_cells_ok or not month_format_ok:
            break

    # Update columns checks pass status
    if column_count_ok:
        checklist["columnCount"] = {"label": "Column Count Check", "status": "pass", "message": "All sheets have the correct number of columns."}
    else:
        fail_remaining("Cannot validate other sheet fields due to column count mismatch.")
        return checklist, False

    if header_blank_ok:
        checklist["headerBlank"] = {"label": "Empty Columns in Header Check", "status": "pass", "message": "All column headers are filled."}
    else:
        fail_remaining("Cannot validate other fields due to empty column header.")
        return checklist, False

    if column_names_ok:
        checklist["columnNames"] = {"label": "Column Names & Order Check", "status": "pass", "message": "All sheet columns and orders match standard."}
    else:
        fail_remaining("Cannot validate other fields due to column name/order mismatch.")
        return checklist, False

    if empty_cells_ok:
        checklist["emptyCells"] = {"label": "Mandatory Column Empty Cell Check", "status": "pass", "message": "All mandatory columns are filled in all sheets."}
    else:
        fail_remaining("Cannot validate month format due to blank mandatory cell error.")
        return checklist, False

    if month_format_ok:
        checklist["monthFormat"] = {"label": "Month Format Validation", "status": "pass", "message": "Month formats correspond to expected schema (e.g. Mar'26)."}
    else:
        return checklist, False

    template_file.seek(0)
    return checklist, True


# =========================================================================
# SPOC MANUAL VALIDATION CHECKS (PHASE 3)
# =========================================================================

def find_column_index(headers, patterns):
    """
    Helper to find the column index matching a regex pattern.
    """
    for pattern in patterns:
        for idx, h in enumerate(headers):
            if h and re.search(pattern, str(h), re.IGNORECASE):
                return idx
    return None


def parse_excel_rows_detail(file_path_or_obj):
    """
    Parses sheets of an Excel workbook and returns a list of row dicts containing payout details.
    """
    try:
        if isinstance(file_path_or_obj, str):
            wb = openpyxl.load_workbook(file_path_or_obj, data_only=True)
        else:
            file_path_or_obj.seek(0)
            wb = openpyxl.load_workbook(file_path_or_obj, data_only=True)
    except Exception as e:
        import logging
        logging.error(f"Error loading workbook: {e}")
        return []

    rows_data = []

    # Patterns to match column headers case-insensitively
    emp_patterns = [r'^emp\s*(?:no|code|id)?$', r'^employee\s*(?:no|code|id)$']
    name_patterns = [r'employee\s*name', r'emp.*name', r'^name$']
    month_patterns = [r'month']
    amount_patterns = [r'amount', r'allowance.*amount', r'shift.*amount', r'copay.*charges', r'total.*deduction']
    bank_patterns = [r'bank\s*a(?:c|c?count)\s*(?:num(?:ber)?)?', r'bank\s*account']
    payout_type_patterns = [r'type.*incentive', r'type.*allowance', r'reason']
    rating_patterns = [r'rating', r'score']

    for sheet_name in wb.sheetnames:
        if sheet_name == "Firstsource Confidential" or wb[sheet_name].sheet_state != 'visible':
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Find header row
        header_row_idx = None
        headers = None
        for r_idx, row in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header_row_idx = r_idx
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                while headers and headers[-1] == "":
                    headers.pop()
                break

        if header_row_idx is None:
            continue

        # Locate indexes
        emp_idx = find_column_index(headers, emp_patterns)
        name_idx = find_column_index(headers, name_patterns)
        month_idx = find_column_index(headers, month_patterns)
        amount_idx = find_column_index(headers, amount_patterns)
        bank_idx = find_column_index(headers, bank_patterns)
        payout_type_idx = find_column_index(headers, payout_type_patterns)
        rating_idx = find_column_index(headers, rating_patterns)

        # Default values if index is missing
        if emp_idx is None:
            emp_idx = 0
        if name_idx is None:
            name_idx = 1 if len(headers) > 1 else 0

        # Validate rows
        for r_idx in range(header_row_idx + 1, len(all_rows)):
            row = all_rows[r_idx]
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            emp_no = str(row[emp_idx]).strip() if emp_idx < len(row) and row[emp_idx] is not None else ""
            emp_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""

            if not emp_no or emp_no.lower() == "none":
                continue

            # Month
            raw_month = row[month_idx] if month_idx is not None and month_idx < len(row) else ""
            if isinstance(raw_month, (datetime, date)):
                month_str = raw_month.strftime("%b'%y")
            else:
                month_str = str(raw_month or "").strip().replace("’", "'")

            # Amount
            raw_amount = row[amount_idx] if amount_idx is not None and amount_idx < len(row) else 0.0
            try:
                amount_val = float(str(raw_amount).replace(",", "").strip()) if raw_amount is not None else 0.0
            except ValueError:
                amount_val = 0.0

            # Bank Account
            bank_val = str(row[bank_idx]).strip() if bank_idx is not None and bank_idx < len(row) and row[bank_idx] is not None else None

            # Payout Type
            payout_val = str(row[payout_type_idx]).strip() if payout_type_idx is not None and payout_type_idx < len(row) and row[payout_type_idx] is not None else sheet_name

            # Rating
            raw_rating = row[rating_idx] if rating_idx is not None and rating_idx < len(row) else ""
            rating_val = str(raw_rating).strip() if raw_rating is not None else ""

            # Frequency (Arrears vs Regular)
            frequency_val = "Arrears" if "arrear" in sheet_name.lower() else "Regular"

            rows_data.append({
                "sheet_name": sheet_name,
                "row_idx": r_idx + 1,
                "emp_no": emp_no,
                "emp_name": emp_name,
                "month": month_str,
                "amount": amount_val,
                "bank_account": bank_val,
                "payout_type": payout_val,
                "rating": rating_val,
                "frequency": frequency_val
            })

    if not isinstance(file_path_or_obj, str):
        file_path_or_obj.seek(0)
    return rows_data


def analyze_and_generate_flags(log_id, file_path_or_obj, doc_code, user_geo):
    """
    Scans the uploaded Excel file, runs duplicate and validation checks,
    and inserts flags into tbl_UploadRowFlags.
    """
    from .db import get_payroll_connection
    from django.conf import settings
    
    current_rows = parse_excel_rows_detail(file_path_or_obj)
    if not current_rows:
        return

    # Delete any existing flags for this LogId first to avoid duplicate flag insertion
    conn = get_payroll_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM tbl_UploadRowFlags WHERE LogId = ?", [log_id])
        conn.commit()
    except Exception as e:
        import logging
        logging.error(f"Error clearing existing flags: {e}")

    # Helper function to extract month from file name
    import re
    def extract_month_from_filename(filename):
        if not filename:
            return ""
        month_regex = re.compile(
            r"(?:^|[^a-zA-Z])(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s'_-]*(\d{2,4})\b",
            re.IGNORECASE
        )
        match = month_regex.search(filename)
        if match:
            m = match.group(1)[:3].capitalize()
            y = match.group(2)
            if len(y) == 4:
                y = y[2:]
            return f"{m}'{y}"
        return ""

    # Fetch original filename and upload date for fallback month extraction
    original_filename = ""
    upload_month = ""
    try:
        cursor.execute("SELECT FileName, CreatedDate FROM tbl_UploadSuccessLog WHERE LogId = ?", [log_id])
        res = cursor.fetchone()
        if res:
            original_filename = res[0] or ""
            if res[1]:
                upload_month = res[1].strftime("%b'%y")
    except Exception as e:
        import logging
        logging.error(f"Error querying filename/upload date from log: {e}")

    # Gather past approved/merged files in the last 12 rolling months
    past_payouts = []
    import datetime as dt
    twelve_months_ago = dt.datetime.now() - dt.timedelta(days=365)
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT LogId, FilePath, FileName, CreatedDate 
            FROM tbl_UploadSuccessLog 
            WHERE IsDeleted = 0 
            AND StateId IN (5, 8, 10) 
            AND UserGeo = ? 
            AND DocumentType = ?
            AND CreatedDate >= ?
            AND LogId != ?
        """, [user_geo, doc_code, twelve_months_ago, log_id])
        past_files = cursor.fetchall()
        
        for p_log_id, p_file_path, p_file_name, p_created_date in past_files:
            rel_path = p_file_path.replace('\\', '/').lstrip('/')
            abs_path = os.path.join(settings.BASE_DIR, rel_path)
            if os.path.exists(abs_path):
                p_rows = parse_excel_rows_detail(abs_path)
                for pr in p_rows:
                    pr["source_file"] = p_file_name
                    pr["source_log_id"] = p_log_id
                    pr["source_date"] = p_created_date.strftime("%d/%b/%Y")
                    past_payouts.append(pr)
    except Exception as e:
        import logging
        logging.error(f"Error loading past payouts: {e}")

    # Get employee master names for Check 4
    employee_master_names = {}
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT EmployeeCode, EmployeeName FROM tbl_EmployeeMaster WHERE DeletedStatus = 0")
        for emp_code, emp_name in cursor.fetchall():
            if emp_code:
                employee_master_names[str(emp_code).strip()] = str(emp_name).strip()
    except Exception as e:
        import logging
        logging.error(f"Error loading employee master names: {e}")

    triggered_flags = {}

    def raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, payout_month, amount, bank_account, flag_type, flag_message):
        key = (sheet_name, row_idx)
        if key in triggered_flags:
            existing = triggered_flags[key]
            triggered_flags[key] = {
                'flag_type': f"{existing['flag_type']}, {flag_type}",
                'flag_message': f"{existing['flag_message']} | {flag_message}"
            }
        else:
            triggered_flags[key] = {
                'flag_type': flag_type,
                'flag_message': flag_message
            }

    # Run Checks for each row in current upload
    for row in current_rows:
        sheet_name = row["sheet_name"]
        row_idx = row["row_idx"]
        emp_no = row["emp_no"]
        emp_name = row["emp_name"]
        month = row["month"]
        amount = row["amount"]
        bank_account = row["bank_account"]
        payout_type = row["payout_type"]
        frequency = row["frequency"]

        # --- Check 1: Duplicate payout check across 12 rolling months of data ---
        for pr in past_payouts:
            if (pr["emp_no"] == emp_no and 
                pr["month"].lower() == month.lower() and 
                pr["frequency"].lower() == frequency.lower() and 
                abs(pr["amount"] - amount) < 0.01 and 
                pr["payout_type"].lower() == payout_type.lower()):
                
                msg = f"Duplicate check: Identical payout already approved in file '{pr['source_file']}' uploaded on {pr['source_date']}."
                raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "DuplicatePayout", msg)
                break

        # --- Check 1b: Duplicate check against rolling ETO team database (SSMS) ---
        try:
            target_month = month
            if not target_month:
                target_month = extract_month_from_filename(original_filename)
                if not target_month:
                    target_month = extract_month_from_filename(os.path.basename(str(file_path_or_obj)))
                    if not target_month:
                        target_month = upload_month

            eto_payout_type = payout_type
            if doc_code == 'IRF':
                eto_payout_type = 'IRF'
            elif doc_code == 'TDCT':
                eto_payout_type = 'TDCT'
            elif doc_code == 'PRL' and payout_type in ('VP Ratings', 'VP Arrears'):
                eto_payout_type = 'PRL'

            cursor.execute("""
                DECLARE @is_dup BIT;
                DECLARE @details VARCHAR(1000);
                EXEC SP_ValidateETODuplicate 
                    @EmpNo = ?, 
                    @EmployeeName = ?, 
                    @PayoutMonth = ?, 
                    @Frequency = ?, 
                    @Amount = ?, 
                    @PayoutType = ?, 
                    @IsDuplicate = @is_dup OUTPUT, 
                    @DuplicateDetails = @details OUTPUT;
                SELECT @is_dup, @details;
            """, [emp_no, emp_name, target_month, frequency, amount, eto_payout_type])
            row_res = cursor.fetchone()
            if row_res:
                is_duplicate = bool(row_res[0])
                dup_details = str(row_res[1] or '')
                if is_duplicate:
                    raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "DuplicatePayout", dup_details)
        except Exception as sp_err:
            import logging
            logging.error(f"Error calling SP_ValidateETODuplicate for {emp_no}: {sp_err}")

        # --- Check 1c: Duplicate payout check within the current Excel file itself ---
        for other_row in current_rows:
            if (other_row["row_idx"] != row_idx and
                other_row["emp_no"] == emp_no and
                other_row["month"].lower() == month.lower() and
                other_row["frequency"].lower() == frequency.lower() and
                abs(other_row["amount"] - amount) < 0.01 and
                other_row["payout_type"].lower() == payout_type.lower()):
                
                msg = f"Duplicate check: Identical payout duplicate found within this Excel file at row {other_row['row_idx']}."
                raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "DuplicatePayout", msg)
                break

        # --- Check 2: Same bank account used across multiple Employee IDs ---
        if bank_account and str(bank_account).strip().lower() != "none" and str(bank_account).strip() != "":
            # Check within the current upload
            for other_row in current_rows:
                if other_row != row and other_row["bank_account"] == bank_account and other_row["emp_no"] != emp_no:
                    msg = f"Bank Account check: Same bank account number '{bank_account}' is used by another employee '{other_row['emp_name']}' (ID: {other_row['emp_no']}) in this file."
                    raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "DuplicateBank", msg)
                    break
            # Check across past uploads (12 months)
            for pr in past_payouts:
                if pr["bank_account"] == bank_account and pr["emp_no"] != emp_no:
                    msg = f"Bank Account check: Same bank account number '{bank_account}' was used by another employee '{pr['emp_name']}' (ID: {pr['emp_no']}) in file '{pr['source_file']}'."
                    raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "DuplicateBank", msg)
                    break

        # --- Check 4: Same employee ID with different employee name ---
        # 4a. Within the current file
        for other_row in current_rows:
            if other_row["emp_no"] == emp_no and other_row["emp_name"].lower().strip() != emp_name.lower().strip():
                msg = f"Name check: Same Employee ID {emp_no} is associated with different employee names ('{emp_name}' vs '{other_row['emp_name']}') in this file."
                raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "NameMismatch", msg)
                break
        # 4b. Against Employee Master
        if emp_no in employee_master_names:
            master_name = employee_master_names[emp_no]
            if master_name.lower().strip() != emp_name.lower().strip():
                msg = f"Name check: Name '{emp_name}' does not match Employee Master name '{master_name}'."
                raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "NameMismatch", msg)
        # 4c. Against past uploads (rolling 12 months)
        for pr in past_payouts:
            if pr["emp_no"] == emp_no and pr["emp_name"].lower().strip() != emp_name.lower().strip():
                msg = f"Name check: Same Employee ID {emp_no} is associated with different employee name '{pr['emp_name']}' in past file '{pr['source_file']}'."
                raise_flag(sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account, "NameMismatch", msg)
                break


    # --- Check 3: Group uploads during the same month (keep highest rating/amount) ---
    try:
        now = dt.datetime.now()
        start_of_month = dt.datetime(now.year, now.month, 1)
        if now.month == 12:
            end_of_month = dt.datetime(now.year + 1, 1, 1)
        else:
            end_of_month = dt.datetime(now.year, now.month + 1, 1)
            
        cursor = conn.cursor()
        cursor.execute("""
            SELECT LogId, FilePath, FileName, CreatedDate 
            FROM tbl_UploadSuccessLog 
            WHERE IsDeleted = 0 
            AND UserGeo = ? 
            AND DocumentType = ?
            AND CreatedDate >= ? AND CreatedDate < ?
            AND LogId != ?
        """, [user_geo, doc_code, start_of_month, end_of_month, log_id])
        same_month_files = cursor.fetchall()
        
        same_month_payouts = []
        for sm_log_id, sm_file_path, sm_file_name, sm_created_date in same_month_files:
            rel_path = sm_file_path.replace('\\', '/').lstrip('/')
            abs_path = os.path.join(settings.BASE_DIR, rel_path)
            if os.path.exists(abs_path):
                sm_rows = parse_excel_rows_detail(abs_path)
                for smr in sm_rows:
                    smr["source_file"] = sm_file_name
                    smr["source_log_id"] = sm_log_id
                    same_month_payouts.append(smr)
                    
        # Group current rows by Employee ID + Payout Month
        from collections import defaultdict
        current_grouped = defaultdict(list)
        for r in current_rows:
            current_grouped[(r["emp_no"], r["month"])].append(r)
            
        # Group same month payouts by Employee ID + Payout Month
        same_month_grouped = defaultdict(list)
        for r in same_month_payouts:
            same_month_grouped[(r["emp_no"], r["month"])].append(r)
            
        # For each employee+month in current upload:
        for (emp_no, month), cur_list in current_grouped.items():
            sm_list = same_month_grouped.get((emp_no, month), [])
            all_items = cur_list + sm_list
            
            if len(all_items) > 1:
                max_amount = max(item["amount"] for item in all_items)
                
                has_diff = False
                first_item = all_items[0]
                for item in all_items:
                    if abs(item["amount"] - first_item["amount"]) > 0.01 or item["rating"] != first_item["rating"]:
                        has_diff = True
                        break
                        
                if has_diff:
                    for crow in cur_list:
                        if crow["amount"] < max_amount:
                            msg = f"Multiple uploads check: Another upload/row for this employee received during the same month contains a higher amount ({max_amount}). This row item is flagged to be declined."
                            raise_flag(crow["sheet_name"], crow["row_idx"], crow["emp_no"], crow["emp_name"], crow["payout_type"], crow["month"], crow["amount"], crow["bank_account"], "MultipleUpload", msg)
    except Exception as e:
        import logging
        logging.error(f"Error running Check 3: {e}")

    # Insert all rows (flagged and clean) into tbl_UploadRowFlags
    try:
        cursor = conn.cursor()
        for row in current_rows:
            sheet_name = row["sheet_name"]
            row_idx = row["row_idx"]
            emp_no = row["emp_no"]
            emp_name = row["emp_name"]
            payout_type = row["payout_type"]
            month = row["month"]
            amount = row["amount"]
            bank_account = row["bank_account"]
            
            key = (sheet_name, row_idx)
            if key in triggered_flags:
                flag_type = triggered_flags[key]['flag_type']
                flag_message = triggered_flags[key]['flag_message']
                flag_status = 'Yes'
                flag_reason = flag_message
                spoc_action = 'Pending'
            else:
                flag_type = ''
                flag_message = ''
                flag_status = 'No'
                flag_reason = ''
                spoc_action = 'No Action'
                
            cursor.execute("""
                INSERT INTO tbl_UploadRowFlags (
                    LogId, SheetName, RowIndex, EmpNo, EmpName, PayoutType, PayoutMonth, Amount, BankAccount, 
                    FlagType, FlagMessage, SpocAction, FlagStatus, FlagReason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                log_id, sheet_name, row_idx, emp_no, emp_name, payout_type, month, amount, bank_account,
                flag_type, flag_message, spoc_action, flag_status, flag_reason
            ])
        conn.commit()
    except Exception as ex:
        import logging
        logging.error(f"Error inserting validation rows to DB: {ex}")
    finally:
        conn.close()

